# -*- coding: utf-8 -*-
import os
import re
from datetime import datetime
from pathlib import Path
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- تحديد المسار الأساسي للشهر الحالي ---
current_month = datetime.now().month

desktop_path = Path.home() / "OneDrive" / "Desktop"
if not desktop_path.exists():
    desktop_path = Path.home() / "Desktop"

BASE_DIR = desktop_path / f"Month_{current_month}" / "Bank-transfers-and-invoices"
EXCEL_REPORT_PATH = BASE_DIR / f"سجل_حوالات_شهر_{current_month}.xlsx"


def sanitize_filename(filename: str) -> str:
    """تنظيف وتجميل اسم الملف وإزالة المسافات والرموز الزائدة"""
    cleaned = re.sub(r'[\\/*?:"<>|]', '', filename)
    cleaned = re.sub(r'\s*[\.-]{2,}\s*', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()


def extract_receipt_data(pdf_path: Path) -> dict:
    """محرك استخراج البيانات للحوالات الفردية والجماعية والمدفوعات الحكومية"""
    data = {
        "file_name_orig": pdf_path.name,
        "beneficiary": "",
        "amount": 0.0,
        "date": "",
        "ref_num": "",
        "trans_type": "حوالة",
        "notes": "",
        "is_multi": False
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

            # 1. استخراج التاريخ
            date_match = re.search(r'(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})', full_text)
            if date_match:
                raw_date = date_match.group(1)
                if "/" in raw_date:
                    parts = raw_date.split("/")
                    if len(parts[2]) == 4:
                        data["date"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        data["date"] = raw_date
                else:
                    data["date"] = raw_date

            # 2. تحديد نوع العملية وقراءة الملاحظات والخدمات الحكومية
            if "مدفوعات حكومية" in full_text or "MOIPAYMENT" in pdf_path.name.upper():
                data["trans_type"] = "سداد حكومي"
                
                service_match = re.search(r'الخدمة\s*\|\s*([\u0600-\u06FF\s]{2,30})', full_text)
                gov_ben = re.search(r'اسم المستفيد\s*\|\s*([A-Za-z0-9\s]{3,35}|[\u0600-\u06FF\s]{3,35})', full_text)
                
                ben_str = ""
                if service_match:
                    ben_str += service_match.group(1).strip()
                if gov_ben:
                    ben_str += f" - {gov_ben.group(1).strip()}" if ben_str else gov_ben.group(1).strip()
                
                data["beneficiary"] = ben_str if ben_str else "خدمات حكومية"

            elif "حوالة محلية" in full_text or "LOCALTRANSFER" in pdf_path.name.upper():
                data["trans_type"] = "حوالة محلية"
            else:
                data["trans_type"] = "تحويل أهلي"

            # 3. استخراج المستفيدين والمبالغ (التعرف على الإيصالات المجمعة)
            beneficiaries = re.findall(r'المستفيد\s*\|\s*([\u0600-\u06FF\s\.-]{3,40})', full_text)
            amounts_raw = re.findall(r'(\d{1,3}(?:,\d{3})*\.\d{2})', full_text)
            
            clean_amounts = []
            for amt in amounts_raw:
                try:
                    val = float(amt.replace(',', ''))
                    if val > 0:
                        clean_amounts.append(val)
                except ValueError:
                    continue

            unique_bens = list(dict.fromkeys([b.strip() for b in beneficiaries if len(b.strip()) > 2]))
            
            if len(unique_bens) > 1:
                data["is_multi"] = True
                first_ben = sanitize_filename(unique_bens[0].split()[0])
                data["beneficiary"] = f"حوالة مجمعة ({len(unique_bens)} مستفيدين) - {first_ben} وآخرون"
                data["amount"] = sum(clean_amounts) if clean_amounts else 0.0
            elif len(unique_bens) == 1 and not data["beneficiary"]:
                data["beneficiary"] = unique_bens[0]
                data["amount"] = clean_amounts[0] if clean_amounts else 0.0
            elif clean_amounts and data["amount"] == 0.0:
                data["amount"] = max(clean_amounts)

            # 4. تنظيف الاسم النهائي للمستفيد
            if data["beneficiary"]:
                data["beneficiary"] = re.split(r'(الرسوم|المبلغ|الحساب|تاريخ|نوع|مرجع|ملاحظات)', data["beneficiary"])[0].strip()
                data["beneficiary"] = sanitize_filename(data["beneficiary"])

            # 5. استخراج الملاحظات المكتوبة إن وجدت
            notes_match = re.search(r'ملاحظات\s*\|\s*([\u0600-\u06FF\s0-9]{2,30})', full_text)
            if notes_match:
                note_val = notes_match.group(1).strip()
                if note_val and note_val not in ["الكل", "المعمد التالي", "فوري"]:
                    data["notes"] = note_val

            # 6. استخراج رقم المرجع
            ref_match = re.search(r'(?:مرجع العملية|مرجع)\s*\|\s*(\d{8,12})', full_text)
            if ref_match:
                data["ref_num"] = ref_match.group(1)

    except Exception as e:
        print(f"[-] خطأ أثناء تحليل {pdf_path.name}: {e}")

    return data


def generate_short_filename(data: dict) -> str:
    """إنشاء مسمى مختصر ونظيف"""
    date_part = f"{data['date']} - " if data["date"] else ""
    ben_part = data["beneficiary"] if data["beneficiary"] else "حوالة"
    amt_part = f"{data['amount']:,.2f}" if data["amount"] > 0 else ""
    note_part = f" ({data['notes']})" if data["notes"] else ""

    if amt_part:
        new_name = f"{date_part}{ben_part}{note_part} - {amt_part}.pdf"
    else:
        new_name = f"{date_part}{ben_part}{note_part}.pdf"

    return sanitize_filename(new_name)


def create_excel_report(records: list, output_path: Path):
    """إنشاء تقرير الإكسيل"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"إيصالات شهر {current_month}"
    ws.views.sheetView[0].rightToLeft = True

    headers = ["التاريخ", "نوع العملية", "المستفيد / الجهة", "المبلغ (ر.س)", "الملاحظات", "رقم المرجع", "اسم الملف"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                         right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'),
                         bottom=Side(style='thin', color='D9D9D9'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    for row_idx, record in enumerate(records, start=2):
        row_data = [
            record["date"],
            record["trans_type"],
            record["beneficiary"],
            record["amount"],
            record["notes"],
            record["ref_num"],
            record["new_file_name"]
        ]
        ws.append(row_data)

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            
            if row_idx % 2 == 0:
                cell.fill = zebra_fill

            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [1, 2, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    total_row = len(records) + 2
    ws.cell(row=total_row, column=3, value="الإجمالي العام").font = Font(bold=True)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="center")
    
    total_cell = ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
    total_cell.font = Font(bold=True, color="1F4E78")
    total_cell.number_format = '#,##0.00'
    total_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output_path)


def process_all_receipts():
    """المعالج المباشر"""
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    pdf_files = list(BASE_DIR.glob("*.pdf")) + list(BASE_DIR.glob("*.PDF"))
    
    if not pdf_files:
        print(f"[!] لا توجد ملفات PDF في المجلد: {BASE_DIR}")
        return

    print(f"[*] جاري معالجة {len(pdf_files)} ملف في المجلد الأساسي...")
    records = []

    for pdf_path in pdf_files:
        data = extract_receipt_data(pdf_path)
        new_filename = generate_short_filename(data)
        
        data["new_file_name"] = new_filename
        new_path = BASE_DIR / new_filename

        if pdf_path.name != new_filename:
            counter = 1
            while new_path.exists() and new_path != pdf_path:
                stem = new_path.stem
                new_path = BASE_DIR / f"{stem}_{counter}.pdf"
                counter += 1

            try:
                pdf_path.rename(new_path)
                data["new_file_name"] = new_path.name
                print(f"[✓] تم تعديل الاسم: {new_path.name}")
            except Exception as e:
                print(f"[-] تعذر إعادة تسمية {pdf_path.name}: {e}")

        records.append(data)

    if records:
        create_excel_report(records, EXCEL_REPORT_PATH)
        print(f"\n[🎉] اكتملت العملية بنجاح! تم تنظيف وتوحيد كافة المسميات.")

if __name__ == "__main__":
    process_all_receipts()